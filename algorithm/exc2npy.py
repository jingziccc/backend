from openpyxl import load_workbook
import torch
import torchvision
from torch import nn
import copy
import numpy as np
import sys
import os

if len(sys.argv) < 2:
    print("Usage: python exc2npy.py <excelfile>")
    sys.exit(1)
# 第一个参数是脚本名称,第二个参数是文件名
excel_file = sys.argv[1]
# 分隔文件名和扩展名
file_name, ext = os.path.splitext(excel_file)
print(file_name, ext)


df = load_workbook(excel_file)
sheet=df.active
a=[]
for i in sheet.iter_rows():
    i_d=[cell.value for cell in i]
    a.append(i_d)
a=np.array(a)
aa = np.reshape(a,(a.shape[0]//18//3,3,18,18))
tbt = torch.tensor(aa)
model = torchvision.models.resnet50(pretrained = True)
model = model
for param in model.parameters():
    param.requires_grad = False
model_last = model.fc.in_features
model.conv1 = nn.Conv2d(3, 64, kernel_size=(3, 3), stride=(1, 1), padding=(3, 3), bias=False)
model.fc = nn.Sequential(nn.Linear(model_last,128),
                        nn.ReLU(inplace = True),
                        nn.Linear(128,70))
loss_intitial = copy.deepcopy(model.cuda())
def test(tbt,model):       
        images=tbt.cuda()
        images = images.float()
        outputs=model(images)
        _,predicted=torch.max(outputs.data,1)
        print(predicted)
model = torch.load(r'./model/model_resnet_1.pt')#加载模型
test(tbt,model)


